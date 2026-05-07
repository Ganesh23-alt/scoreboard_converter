import json
import re
from datetime import datetime, date
from openpyxl import load_workbook

INPUT_FILE = "scoreboard_test.xlsx"
OUTPUT_FILE = "output.json"
SHEET_NAME = "SCOREBOARD"

# ----------------------------
# Helpers
# ----------------------------

def normalize_key(name):
    key = str(name).lower()
    key = re.sub(r"[^a-z0-9]+", "_", key)
    key = re.sub(r"_+", "_", key)
    return key.strip("_")


def clean_value(val):
    if val is None:
        return None

    if isinstance(val, (datetime, date)):
        return val.isoformat()

    if isinstance(val, str):
        val = val.strip()
        return val if val != "" else None

    if isinstance(val, float):
        return round(val, 2)

    return val


# ----------------------------
# Merged cell handling (OPTIMIZED)
# ----------------------------

def build_merged_map(ws):
    merged_map = {}

    for merged in ws.merged_cells.ranges:
        value = ws.cell(merged.min_row, merged.min_col).value

        for r in range(merged.min_row, merged.max_row + 1):
            for c in range(merged.min_col, merged.max_col + 1):
                merged_map[(r, c)] = value

    return merged_map


def get_value(ws, merged_map, row, col):
    return ws.cell(row=row, column=col).value or merged_map.get((row, col))


# ----------------------------
# Core transformation
# ----------------------------

def extract_metrics(ws):
    merged_map = build_merged_map(ws)

    HEADER_ROW = 2
    FOCUS_ROW = 3
    SOURCE_ROW = 4
    ROLE_ROW = 5
    TARGET_ROW = 7
    DATA_START_ROW = 8

    max_col = ws.max_column
    max_row = ws.max_row
    date_col = 1

    headers = {}
    focuses = {}
    sources = {}
    roles = {}
    targets = {}

    valid_cols = []

    # ----------------------------
    # Read column metadata
    # ----------------------------
    for col in range(2, max_col + 1):
        header = clean_value(get_value(ws, merged_map, HEADER_ROW, col))

        if not header:
            continue

        valid_cols.append(col)

        headers[col] = header
        focuses[col] = clean_value(get_value(ws, merged_map, FOCUS_ROW, col))
        sources[col] = clean_value(get_value(ws, merged_map, SOURCE_ROW, col))
        roles[col] = clean_value(get_value(ws, merged_map, ROLE_ROW, col))
        targets[col] = clean_value(get_value(ws, merged_map, TARGET_ROW, col))

    # ----------------------------
    # Read row data
    # ----------------------------
    data_rows = []

    for row in range(DATA_START_ROW, max_row + 1):
        date_val = clean_value(get_value(ws, merged_map, row, date_col))

        if not date_val:
            continue

        row_data = {"date": date_val}

        for col in valid_cols:
            row_data[col] = clean_value(get_value(ws, merged_map, row, col))

        data_rows.append(row_data)

    # ----------------------------
    # Build metrics
    # ----------------------------
    metrics = []

    for col in valid_cols:
        display_name = headers[col]

        metric = {
            "metric_key": normalize_key(display_name),
            "display_name": display_name,
            "focus": focuses.get(col),
            "source": sources.get(col),
            "role": roles.get(col),
            "target": targets.get(col),
            "data": []
        }

        for row in data_rows:
            date_val = row.get("date")
            metric_val = row.get(col)

            if date_val and metric_val is not None:
                metric["data"].append({
                    "date": str(date_val),
                    "value": metric_val
                })

        if metric["data"]:
            metrics.append(metric)

    return metrics


# ----------------------------
# Output builder
# ----------------------------

def build_output(metrics):
    return {
        "sheet_name": SHEET_NAME,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "metrics": metrics
    }


# ----------------------------
# Main
# ----------------------------

def main():
    wb = load_workbook(INPUT_FILE, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found.")

    ws = wb[SHEET_NAME]

    metrics = extract_metrics(ws)
    output = build_output(metrics)

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print(f"Conversion complete → {OUTPUT_FILE}")


if __name__ == "__main__":
    main()